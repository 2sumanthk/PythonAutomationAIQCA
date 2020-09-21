import openpyxl


class ReadWriteExcel:
    work_book = None
    file_path = None

    def __init__(self, path):
        global work_book
        global file_path

        file_path = path

        work_book = openpyxl.load_workbook(file_path)

    def read_by_col_index(self, sheetname, rowindex, colindex):
        sheet = work_book[sheetname]
        return sheet.cell(rowindex, colindex).value

    def read_by_col_name(self, sheetname, rowindex, colname):
        sheet = work_book[sheetname]
        colindex = 1
        while sheet.cell(row=1, column=colindex).value != '':
            if colname == sheet.cell(row=1, column=colindex).value:
                break
            colindex = colindex + 1
        return sheet.cell(rowindex, colindex).value

    def get_cell_value(self,sheetname, rowindex, colindex):
        sheet = work_book[sheetname]
        return sheet.cell(rowindex, colindex).value

    def get_row_count(self, sheetname):
        sheet = work_book[sheetname]
        return sheet.max_row

    def get_col_count(self, sheetname):
        sheet = work_book[sheetname]
        return sheet.max_column

    def write_by_col_index(self, sheetname, rowindex, colindex, value):
        sheet = work_book[sheetname]
        sheet.cell(row=rowindex, column=colindex).value = value
        work_book.save(file_path)
